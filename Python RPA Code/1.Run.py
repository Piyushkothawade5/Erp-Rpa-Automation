import pyautogui
import time
import keyboard
import sys
import pyperclip
import os
import threading
from datetime import datetime
from openpyxl import Workbook, load_workbook

# ================= SAFETY =================
pyautogui.FAILSAFE = True

# ================= FILES =================
FILE_PATH = "items.txt"
PROGRESS_FILE = "progress_items.txt"
RESULT_FILE = "item_validation_log.xlsx"

# ================= USER INPUT =================
TARGET_DATE = "01/04/2026"
BATCH_SIZE = 20

# ================= COORDINATES =================
INSERT_BTN = (83, 342)
SAVE_BTN = (134, 64)
SAVE_OK = (450, 239)
APPROVE_BTN = (430, 66)
ENTERED_BY_BTN = (61, 118)
APPROVED_BY_BTN = (44, 135)
OK_BTN = (340, 321)
ERROR_OK_BTN = (383, 239)   # invalid popup OK
NEW_FORM_BTN = (56, 67)

# ================= TIMINGS =================
START_DELAY = 5
DELAY = 0.5
ENTER_DELAY = 0.25
POPUP_WAIT_SECONDS = 1
POPUP_WATCH_INTERVAL = 0.6

# ================= STOP / STATE =================
current_index = 0
popup_stop_event = threading.Event()

# ================= PROGRESS =================
def save_progress(index):
    with open(PROGRESS_FILE, "w") as f:
        f.write(str(index))

def load_progress():
    try:
        with open(PROGRESS_FILE, "r") as f:
            return int(f.read().strip())
    except:
        return 0

def check_stop():
    if keyboard.is_pressed('esc'):
        print("\n🛑 STOP pressed. Saving progress...")
        save_progress(current_index)
        safe_save_workbook()
        popup_stop_event.set()
        sys.exit()

# ================= EXCEL LOGGING =================
HEADERS = ["Input_SrNo", "Valid_SrNo", "ItemCode", "Location", "Qty", "Status", "Time"]

wb = None
ws = None
dirty_wb = False

def init_workbook():
    global wb, ws
    if os.path.exists(RESULT_FILE):
        wb = load_workbook(RESULT_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Log"
        ws.append(HEADERS)
        wb.save(RESULT_FILE)

def load_last_valid_srno():
    if not os.path.exists(RESULT_FILE):
        return 0

    wb_local = load_workbook(RESULT_FILE)
    ws_local = wb_local.active

    last = 0
    for r in range(ws_local.max_row, 1, -1):
        v = ws_local.cell(row=r, column=2).value
        if v is not None and str(v).strip() != "":
            try:
                last = int(v)
                break
            except:
                pass
    return last

def log_row(input_sr, valid_sr, item, loc, qty, status):
    global dirty_wb
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([input_sr, valid_sr, item, loc, qty, status, now])
    dirty_wb = True

def safe_save_workbook():
    global dirty_wb
    if dirty_wb:
        wb.save(RESULT_FILE)
        dirty_wb = False

# ================= POPUP HANDLING =================
def clear_general_popup_once():
    """
    Closes the general popup if it is visible.
    Returns True if clicked, else False.
    """
    try:
        location = pyautogui.locateOnScreen('thanks_btn.png', confidence=0.8, grayscale=True)
        if location:
            pyautogui.click(pyautogui.center(location))
            time.sleep(0.5)
            return True
    except:
        pass
    return False

def clear_general_popups(max_rounds=3):
    """
    Clears general popups repeatedly for a short burst.
    Useful before and after important actions.
    """
    for _ in range(max_rounds):
        check_stop()
        if not clear_general_popup_once():
            break

def popup_watcher():
    """
    Background watcher that keeps clearing the general popup while the script runs.
    """
    while not popup_stop_event.is_set():
        try:
            cleared = clear_general_popup_once()
            time.sleep(0.15 if cleared else POPUP_WATCH_INTERVAL)
        except:
            time.sleep(POPUP_WATCH_INTERVAL)

def invalid_popup_present():
    try:
        location = pyautogui.locateOnScreen('invalid_item.png', confidence=0.8, grayscale=True)
        return bool(location)
    except:
        return False

def handle_invalid_popup(item):
    if invalid_popup_present():
        print(f"❌ Invalid Item: {item}")
        pyautogui.click(ERROR_OK_BTN)
        time.sleep(0.8)
        clear_general_popups()
        return True
    return False

def wait_for_invalid_popup(item, timeout=POPUP_WAIT_SECONDS):
    end_time = time.time() + timeout
    while time.time() < end_time:
        check_stop()
        if handle_invalid_popup(item):
            return True
        clear_general_popups(max_rounds=1)
        time.sleep(0.15)
    return False

def safe_action(action, after_delay=0.0, post_popup_checks=2):
    """
    Wrap any pyautogui action with popup clearing before and after.
    """
    clear_general_popups()
    action()
    if after_delay:
        time.sleep(after_delay)
    for _ in range(post_popup_checks):
        check_stop()
        clear_general_popups(max_rounds=1)

# ================= LOAD DATA =================
data = []
with open(FILE_PATH, "r") as file:
    for line in file:
        if line.strip():
            item, loc, qty = line.strip().split(",")
            data.append((item.strip(), loc.strip(), qty.strip()))

start_index = load_progress()
print(f"📄 Total Rows: {len(data)}")
print(f"🔁 Resuming from row {start_index + 1}")

init_workbook()
valid_sr_counter = load_last_valid_srno()

time.sleep(START_DELAY)

popup_thread = threading.Thread(target=popup_watcher, daemon=True)
popup_thread.start()

# ================= PROCESS ONE ENTRY =================
def process_entry(item, loc, qty, skip_insert):
    check_stop()
    clear_general_popups()

    if not skip_insert:
        safe_action(lambda: pyautogui.click(INSERT_BTN), after_delay=DELAY)

    # Paste item code
    pyperclip.copy(item)
    time.sleep(0.15)
    safe_action(lambda: pyautogui.hotkey('ctrl', 'v'), after_delay=0.25)

    # Wait for invalid popup BEFORE moving on
    if wait_for_invalid_popup(item):
        return "invalid"

    # Enter x4, but stop instantly if invalid popup appears
    for _ in range(4):
        check_stop()
        if handle_invalid_popup(item):
            return "invalid"
        safe_action(lambda: pyautogui.press('enter'), after_delay=ENTER_DELAY)
        if handle_invalid_popup(item):
            return "invalid"

    # Location
    pyperclip.copy(loc)
    time.sleep(0.15)
    safe_action(lambda: pyautogui.hotkey('ctrl', 'v'), after_delay=DELAY)

    # Qty
    safe_action(lambda: pyautogui.press('enter'), after_delay=0.35)
    pyperclip.copy(qty)
    time.sleep(0.15)
    safe_action(lambda: pyautogui.hotkey('ctrl', 'v'), after_delay=DELAY)

    return "success"

# ================= SAVE & APPROVE =================
def save_and_approve():
    clear_general_popups()
    safe_action(lambda: pyautogui.click(SAVE_BTN))
    safe_action(lambda: pyautogui.click(SAVE_OK), after_delay=2)

    safe_action(lambda: pyautogui.click(APPROVE_BTN), after_delay=DELAY)

    safe_action(lambda: pyautogui.doubleClick(ENTERED_BY_BTN), after_delay=DELAY)
    safe_action(lambda: pyautogui.click(OK_BTN), after_delay=DELAY)

    safe_action(lambda: pyautogui.click(APPROVE_BTN), after_delay=DELAY)

    safe_action(lambda: pyautogui.doubleClick(APPROVED_BY_BTN), after_delay=DELAY)
    safe_action(lambda: pyautogui.click(OK_BTN), after_delay=2)

# ================= NEW TRANSACTION =================
def new_transaction():
    safe_action(lambda: pyautogui.click(NEW_FORM_BTN), after_delay=2)
    pyperclip.copy(TARGET_DATE)
    time.sleep(0.15)
    safe_action(lambda: pyautogui.hotkey('ctrl', 'v'))

# ================= MAIN LOOP =================
count = 0
skip_insert_next = False

for i in range(start_index, len(data)):
    current_index = i
    item, loc, qty = data[i]
    input_sr = i + 1

    check_stop()
    print(f"➡ Processing {input_sr}: {item}")

    result = process_entry(item, loc, qty, skip_insert_next)

    if result == "invalid":
        log_row(input_sr, None, item, loc, qty, "INVALID")
        safe_save_workbook()

        save_progress(i + 1)
        skip_insert_next = True
        continue

    valid_sr_counter += 1
    log_row(input_sr, valid_sr_counter, item, loc, qty, "VALID")

    count += 1
    save_progress(i + 1)
    skip_insert_next = False

    if (input_sr % 10) == 0:
        safe_save_workbook()

    if count == BATCH_SIZE:
        print("💾 Batch complete → Saving & Approving")
        safe_save_workbook()
        save_and_approve()
        new_transaction()
        count = 0
        skip_insert_next = False

# ================= FINAL SAVE =================
print("✅ Final Save...")
safe_save_workbook()
save_and_approve()
save_progress(0)
popup_stop_event.set()

print(f"🎉 DONE. Excel saved as: {RESULT_FILE}")